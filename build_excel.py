#!/usr/bin/env python3
"""Build detailed XLSX itinerary for USA + Hawaii 20-day trip."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------- Styles ----------
H_FILL = PatternFill("solid", fgColor="1F4E78")
H_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"))
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT = Font(bold=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Sheet 1 — Overview
# ============================================================
ws = wb.active
ws.title = "Overview"
ws["A1"] = "🌴 USA + HAVAJ 2027 — 20denní cesta"
ws["A1"].font = Font(bold=True, size=18, color="1F4E78")
ws.merge_cells("A1:E1")

ws["A3"] = "Trasa"
ws["B3"] = "Praha → Los Angeles → Oahu → Maui → Big Island → Kauai → Miami → Praha"
ws["A4"] = "Délka"
ws["B4"] = "20 dní / 18 nocí"
ws["A5"] = "Termín"
ws["B5"] = "TBD (doporučení: duben–červen nebo září–říjen)"
ws["A6"] = "Cestující"
ws["B6"] = "1 osoba (rozpočet) — pro 2 vynásob letenky/aktivity x2, hotely beze změny"
ws["A7"] = "Celkový rozpočet (orientačně, 1 os.)"
ws["B7"] = "~12 700 EUR (střední úroveň)"
ws["B7"].font = Font(bold=True, color="C00000")

for r in range(3, 8):
    ws[f"A{r}"].font = Font(bold=True)

ws["A9"] = "📅 Souhrn nocí"
ws["A9"].font = Font(bold=True, size=13)

ws.append([])
ws.append(["Destinace", "Noci", "Hotel kategorie", "Indikativní cena/noc (EUR)"])
style_header(ws, ws.max_row, 4)

stays = [
    ("Los Angeles — Santa Monica", 1, "4* boutique", "180–280"),
    ("Oahu — Waikiki", 4, "4* resort, ocean view", "220–340"),
    ("Maui — Ka'anapali", 4, "4* beach resort", "320–480"),
    ("Big Island — Volcano + Kona", 3, "Lodge + beach hotel", "180–280"),
    ("Kauai — Princeville/Poipu", 4, "4* resort", "280–400"),
    ("Miami — South Beach", 2, "4* Art Deco hotel", "240–380"),
    ("Letadlo (HNL→MIA noc 17)", 0, "—", "—"),
]
for s in stays:
    ws.append(s)

# Total nights
last = ws.max_row
ws.append(["CELKEM NOCÍ:", sum(s[1] for s in stays), "", ""])
ws.cell(row=ws.max_row, column=1).font = TOTAL_FONT
ws.cell(row=ws.max_row, column=1).fill = TOTAL_FILL
ws.cell(row=ws.max_row, column=2).font = TOTAL_FONT
ws.cell(row=ws.max_row, column=2).fill = TOTAL_FILL

auto_width(ws, [38, 8, 28, 28, 10])

# ============================================================
# Sheet 2 — Day-by-day itinerary
# ============================================================
ws2 = wb.create_sheet("Itinerář dne po dni")
headers = ["Den", "Datum (Pá start)", "Lokalita", "Program", "Doprava",
           "Nocleh", "Aktivity / rezervace"]
ws2.append(headers)
style_header(ws2, 1, len(headers))

days = [
    (1,  "Pá",  "Praha → Los Angeles",
     "Ranní odlet PRG. Přílet LAX odpol. Santa Monica Pier, večeře u oceánu, brzy spát.",
     "PRG→LAX (LH/BA/AF, ~12h s 1× stop)",
     "LA — Santa Monica",
     "—"),
    (2,  "So",  "LA → Honolulu",
     "Ranní let do Honolulu (~6h, -3h posun). Check-in Waikiki, pláž, sunset Diamond Head pohled.",
     "LAX→HNL (Hawaiian / Delta / United)",
     "Oahu / Waikiki",
     "Mai tai u Duke's"),
    (3,  "Ne",  "Oahu — historie",
     "Sunrise Diamond Head trek (~1.5h). Pearl Harbor + USS Arizona Memorial. Odpol. Waikiki.",
     "Uber",
     "Oahu / Waikiki",
     "**Pearl Harbor — REZERVOVAT** (free tickets uvolňuje recreation.gov)"),
    (4,  "Po",  "Oahu — North Shore",
     "Road trip: Haleiwa town, želvy Laniakea Beach, Dole Plantation, sunset Sunset/Banzai Beach.",
     "Půjčené auto (1 den)",
     "Oahu / Waikiki",
     "Shrimp truck Giovanni's"),
    (5,  "Út",  "Oahu — Kualoa + snorkel",
     "Hanauma Bay snorkeling DOPOL (rezervace!) NEBO Kualoa Ranch — Jurský park tour.",
     "Uber / půjč. auto",
     "Oahu / Waikiki",
     "**Hanauma Bay REZERVACE 2 dny předem**"),
    (6,  "St",  "Oahu → Maui",
     "Dopol. Lanikai Beach + Pillbox hike, polední hop HNL→OGG (~40min), check-in Ka'anapali, sunset Black Rock.",
     "HNL→OGG (Hawaiian / Southwest)",
     "Maui / Ka'anapali",
     "Vyzvednout auto na OGG"),
    (7,  "Čt",  "Maui — Road to Hana",
     "Celodenní road trip: vodopády Twin Falls, Wailua, bambusový les Pipiwai, černá pláž Wai'anapanapa.",
     "Auto (celý den)",
     "Maui / Ka'anapali",
     "Start v 6:30! Halfway to Hana banana bread."),
    (8,  "Pá",  "Maui — Haleakala",
     "Sunrise Haleakala (start 3 AM, permit nutný). Odpol. relax Wailea Beach, večer Lahaina front street.",
     "Auto",
     "Maui / Ka'anapali",
     "**Haleakala sunrise permit — REZERVACE 60 dní předem (recreation.gov)**"),
    (9,  "So",  "Maui — Molokini + Lu'au",
     "Ranní katamarán Molokini Crater (snorkeling), odpoledne masáž/bazén, večer Old Lahaina Lu'au.",
     "Vlastní auto",
     "Maui / Ka'anapali",
     "**Old Lahaina Lu'au REZERVACE 30+ dní předem**"),
    (10, "Ne",  "Maui → Big Island",
     "Ranní hop OGG→KOA (~50min). Kona coffee farm tour, sunset Kailua-Kona pier, prejezd do Volcano Village.",
     "OGG→KOA (Hawaiian)",
     "Big Island / Volcano Village",
     "Auto na KOA letišti"),
    (11, "Po",  "Big Island — Volcanoes NP",
     "Celý den Hawaii Volcanoes NP: Kilauea Iki trek, Crater Rim Drive, Thurston Lava Tube, večer červená záře magmatu.",
     "Vlastní auto",
     "Big Island / Volcano Village",
     "Vstup NP $30/auto, platí 7 dní"),
    (12, "Út",  "Big Island — Mauna Kea",
     "Dopol. černá pláž Punaluu (želvy!), zelená pláž Papakōlea. Večer Mauna Kea hvězdárna 4200m + stargazing.",
     "Vlastní auto (4WD pro Mauna Kea!)",
     "Big Island / Kona",
     "**Mauna Kea tour 100–200 USD** (nejlepší s průvodcem — výškové oblečení v ceně)"),
    (13, "St",  "Big Island → Kauai",
     "Dopol. Hapuna Beach, polední let KOA→LIH přes HNL (~3h s přesedáním). Hanalei Bay sunset.",
     "KOA→HNL→LIH (Hawaiian)",
     "Kauai / Princeville",
     "Vrátit auto na KOA, vyzvednout LIH"),
    (14, "Čt",  "Kauai — Na Pali Coast",
     "Celodenní boat tour Na Pali Coast (Holo Holo Charters / Captain Andy's) — útesy, jeskyně, delfíni.",
     "Loď + auto",
     "Kauai / Princeville",
     "**Na Pali boat tour REZERVACE** 250 USD"),
    (15, "Pá",  "Kauai — Waimea Canyon",
     "Dopol. Waimea Canyon (Kalalau Lookout, Pu'u Hinahina), odpol. Polihale Beach západ slunce.",
     "Vlastní auto",
     "Kauai / Poipu",
     "Přejezd na jih ostrova"),
    (16, "So",  "Kauai — Helikoptéra",
     "Ranní helikoptérová tour celého ostrova (Blue Hawaiian / Jack Harter, 50–60 min). Odpol. Poipu Beach, želvy.",
     "Helikoptéra + auto",
     "Kauai / Poipu",
     "**Helikoptéra REZERVACE — 350 USD**"),
    (17, "Ne",  "Kauai → Miami",
     "Dopol. Spouting Horn, balení. Odpol. let LIH→LAX→MIA (overnight, ~12h s layoverem, +6h posun).",
     "LIH→LAX→MIA (Alaska/AA)",
     "Letadlo",
     "Vyspat se v letadle!"),
    (18, "Po",  "Miami",
     "Příl. ráno. Check-in South Beach. Ocean Drive Art Deco, Lincoln Road, odpol. pláž, večer Little Havana (Versailles, Ball & Chain).",
     "Uber",
     "Miami / South Beach",
     "Cuban coffee + cigar"),
    (19, "Út",  "Miami → Praha",
     "Dopol. Everglades airboat (Sawgrass Recreation Park), odpol. Wynwood Walls street art. Večerní let MIA→PRG.",
     "Auto/Uber + MIA→PRG (LH/AF/KL)",
     "Letadlo",
     "**Everglades airboat tour 30 USD**"),
    (20, "St",  "Přílet Praha",
     "Přílet ráno. Welcome home! ✅",
     "—",
     "Doma",
     "—"),
]

for d in days:
    ws2.append(d)

for r in range(2, ws2.max_row + 1):
    for c in range(1, len(headers) + 1):
        cell = ws2.cell(row=r, column=c)
        cell.alignment = WRAP
        cell.border = BORDER
    ws2.row_dimensions[r].height = 70

auto_width(ws2, [5, 14, 24, 65, 32, 26, 42])

# ============================================================
# Sheet 3 — Flights
# ============================================================
ws3 = wb.create_sheet("Letenky")
ws3.append(["Segment", "Trasa", "Dopravce (návrh)", "Délka", "Třída", "Cena (EUR)", "Pozn."])
style_header(ws3, 1, 7)

flights = [
    ("Long-haul outbound", "PRG → LAX (přes FRA/LHR/CDG)", "Lufthansa / British / Air France", "~12 h s 1× přestup", "Economy", 700, "Multi-city / open-jaw rezervace"),
    ("Pacific 1", "LAX → HNL", "Hawaiian / Delta / United", "~6 h přímý", "Economy", 380, ""),
    ("Hawaii hop 1", "HNL → OGG (Maui)", "Hawaiian / Southwest", "~40 min", "Economy", 95, "Bag fee zvlášť"),
    ("Hawaii hop 2", "OGG → KOA (Big Island)", "Hawaiian / Southwest", "~50 min", "Economy", 110, ""),
    ("Hawaii hop 3", "KOA → LIH (Kauai) přes HNL", "Hawaiian", "~3 h s přestupem", "Economy", 180, ""),
    ("Pacific 2 (transit)", "LIH → LAX → MIA (overnight)", "Alaska + American", "~13 h s layoverem", "Economy", 580, "Vybrat rozumný layover 2–3 h"),
    ("Long-haul return", "MIA → PRG (přes FRA/CDG/AMS)", "Lufthansa / Air France / KLM", "~12 h s 1× přestup", "Economy", 650, ""),
]
total_flights = 0
for f in flights:
    ws3.append(f)
    total_flights += f[5]

ws3.append(["", "", "", "", "CELKEM LETENKY:", total_flights, ""])
last = ws3.max_row
for c in range(1, 8):
    ws3.cell(row=last, column=c).fill = TOTAL_FILL
    ws3.cell(row=last, column=c).font = TOTAL_FONT

for r in range(2, ws3.max_row + 1):
    for c in range(1, 8):
        ws3.cell(row=r, column=c).alignment = WRAP
        ws3.cell(row=r, column=c).border = BORDER

auto_width(ws3, [22, 32, 32, 18, 12, 14, 32])

# ============================================================
# Sheet 4 — Accommodation
# ============================================================
ws4 = wb.create_sheet("Ubytování")
ws4.append(["Lokalita", "Hotel (návrh)", "Noci", "Cena/noc (EUR)", "Celkem (EUR)", "Booking link",
            "Pozn."])
style_header(ws4, 1, 7)

hotels = [
    ("LA — Santa Monica", "Shore Hotel / Hotel Casa del Mar", 1, 240, 240,
     "https://www.booking.com/searchresults.html?ss=Santa+Monica",
     "Walking distance to pier"),
    ("Oahu — Waikiki", "The Royal Hawaiian (Pink Palace) / Halekulani", 4, 320, 1280,
     "https://www.booking.com/searchresults.html?ss=Waikiki",
     "Ocean view nutný!"),
    ("Maui — Ka'anapali", "Sheraton Maui / Hyatt Regency Maui", 4, 380, 1520,
     "https://www.booking.com/searchresults.html?ss=Kaanapali",
     "Black Rock přímo u hotelu"),
    ("Big Island — Volcano Village", "Volcano Village Lodge", 1, 220, 220,
     "https://www.booking.com/searchresults.html?ss=Volcano+Village+Hawaii",
     "Klid + blízko NP"),
    ("Big Island — Kona", "Royal Kona Resort / Sheraton Kona", 2, 280, 560,
     "https://www.booking.com/searchresults.html?ss=Kona+Hawaii",
     ""),
    ("Kauai — Princeville", "1 Hotel Hanalei Bay / Princeville Resort", 2, 380, 760,
     "https://www.booking.com/searchresults.html?ss=Princeville+Kauai",
     "Sever ostrova"),
    ("Kauai — Poipu", "Grand Hyatt Kauai / Koa Kea", 2, 360, 720,
     "https://www.booking.com/searchresults.html?ss=Poipu+Kauai",
     "Jih ostrova"),
    ("Miami — South Beach", "The Betsy / Marriott Stanton South Beach", 2, 290, 580,
     "https://www.booking.com/searchresults.html?ss=South+Beach+Miami",
     "Walking k Ocean Drive"),
]
total_hotels = 0
for h in hotels:
    ws4.append(h)
    total_hotels += h[4]

ws4.append(["", "", sum(h[2] for h in hotels), "CELKEM UBYTOVÁNÍ:", total_hotels, "", ""])
last = ws4.max_row
for c in range(1, 8):
    ws4.cell(row=last, column=c).fill = TOTAL_FILL
    ws4.cell(row=last, column=c).font = TOTAL_FONT

# Make booking links clickable
for r in range(2, ws4.max_row + 1):
    cell = ws4.cell(row=r, column=6)
    if cell.value and cell.value.startswith("http"):
        cell.hyperlink = cell.value
        cell.font = Font(color="0563C1", underline="single")
    for c in range(1, 8):
        ws4.cell(row=r, column=c).alignment = WRAP
        ws4.cell(row=r, column=c).border = BORDER

auto_width(ws4, [26, 38, 6, 14, 14, 50, 28])

# ============================================================
# Sheet 5 — Activities
# ============================================================
ws5 = wb.create_sheet("Aktivity")
ws5.append(["Den", "Ostrov / místo", "Aktivita", "Cena (EUR)", "Délka", "Rezervace", "Link"])
style_header(ws5, 1, 7)

acts = [
    (3, "Oahu", "Pearl Harbor + USS Arizona", 0, "3 h", "ANO — recreation.gov, free tickets",
     "https://www.recreation.gov/ticket/facility/233338"),
    (3, "Oahu", "Diamond Head sunrise hike", 5, "1.5 h", "ANO — go-hawaii reservation",
     "https://gostateparks.hawaii.gov/diamondhead"),
    (5, "Oahu", "Hanauma Bay snorkeling", 22, "4 h", "ANO — 2 dny předem online",
     "https://pros8.hnl.info/hanauma-bay"),
    (5, "Oahu", "Kualoa Ranch — Jurassic Adventure", 130, "2 h", "Doporučeno",
     "https://www.kualoa.com/"),
    (7, "Maui", "Road to Hana (self-drive)", 30, "celý den", "Ne (auto a tank)",
     ""),
    (8, "Maui", "Haleakala sunrise permit", 12, "5 h s jízdou", "POVINNĚ — 60 dní předem",
     "https://www.recreation.gov/timed-entry/10084665"),
    (9, "Maui", "Molokini Crater snorkel cruise", 120, "5 h", "ANO — Pacific Whale Foundation",
     "https://www.pacificwhale.org/"),
    (9, "Maui", "Old Lahaina Lu'au", 170, "3 h", "POVINNĚ — 30+ dní předem",
     "https://www.oldlahainaluau.com/"),
    (11, "Big Island", "Hawaii Volcanoes NP entrance", 28, "celý den", "Ne",
     "https://www.nps.gov/havo/"),
    (12, "Big Island", "Mauna Kea Summit + Stars tour", 240, "8 h", "ANO — vč. teplého oblečení",
     "https://www.maunakea.com/"),
    (12, "Big Island", "Punaluu Black Sand + Papakolea Green Sand", 0, "1/2 dne", "Ne",
     ""),
    (14, "Kauai", "Na Pali Coast boat tour", 230, "5 h", "POVINNĚ",
     "https://www.holoholocharters.com/"),
    (15, "Kauai", "Waimea Canyon (self-drive)", 0, "1/2 dne", "Ne",
     ""),
    (16, "Kauai", "Helikoptéra Kauai (Blue Hawaiian)", 320, "1 h", "POVINNĚ",
     "https://www.bluehawaiian.com/kauai/"),
    (18, "Miami", "Wynwood Walls", 12, "2 h", "Ne",
     "https://thewynwoodwalls.com/"),
    (19, "Miami", "Everglades airboat tour", 30, "2 h", "Doporučeno",
     "https://www.sawgrassrecreation.com/"),
]
total_acts = 0
for a in acts:
    ws5.append(a)
    total_acts += a[3]

ws5.append(["", "", "CELKEM AKTIVITY:", total_acts, "", "", ""])
last = ws5.max_row
for c in range(1, 8):
    ws5.cell(row=last, column=c).fill = TOTAL_FILL
    ws5.cell(row=last, column=c).font = TOTAL_FONT

for r in range(2, ws5.max_row + 1):
    cell = ws5.cell(row=r, column=7)
    if cell.value and cell.value.startswith("http"):
        cell.hyperlink = cell.value
        cell.font = Font(color="0563C1", underline="single")
    for c in range(1, 8):
        ws5.cell(row=r, column=c).alignment = WRAP
        ws5.cell(row=r, column=c).border = BORDER

auto_width(ws5, [5, 14, 36, 12, 14, 32, 50])

# ============================================================
# Sheet 6 — Car rental
# ============================================================
ws6 = wb.create_sheet("Auto")
ws6.append(["Lokalita", "Dny", "Cena/den (EUR)", "Celkem (EUR)", "Pozn."])
style_header(ws6, 1, 5)

cars = [
    ("Oahu (na 2 dny — North Shore, Kualoa)", 2, 65, 130, "Volitelné — Uber stačí"),
    ("Maui (po celou dobu)", 4, 70, 280, "POVINNÉ"),
    ("Big Island (po celou dobu)", 3, 75, 225, "POVINNÉ, 4WD pro Mauna Kea"),
    ("Kauai (po celou dobu)", 4, 70, 280, "POVINNÉ"),
    ("Miami (jen Everglades)", 1, 60, 60, "Volitelné"),
]
total_car = 0
for c in cars:
    ws6.append(c)
    total_car += c[3]

ws6.append(["CELKEM AUTO:", "", "", total_car, ""])
last = ws6.max_row
for col in range(1, 6):
    ws6.cell(row=last, column=col).fill = TOTAL_FILL
    ws6.cell(row=last, column=col).font = TOTAL_FONT

for r in range(2, ws6.max_row + 1):
    for c in range(1, 6):
        ws6.cell(row=r, column=c).alignment = WRAP
        ws6.cell(row=r, column=c).border = BORDER

auto_width(ws6, [38, 8, 16, 14, 30])

# ============================================================
# Sheet 7 — Budget summary
# ============================================================
ws7 = wb.create_sheet("Rozpočet")
ws7.append(["Kategorie", "Cena (EUR, 1 os.)", "% z celku", "Pozn."])
style_header(ws7, 1, 4)

food_per_day = 75
food_total = food_per_day * 19  # 19 days

budget = [
    ("Letenky (7 segmentů)", total_flights, ""),
    ("Ubytování (18 nocí)", total_hotels, ""),
    ("Půjčovny aut", total_car, ""),
    ("Aktivity (atrakce, tours)", total_acts, ""),
    ("Jídlo (~75 EUR/den × 19 dní)", food_total, "Havaj a Miami jsou drahé"),
    ("Uber / interní doprava", 250, ""),
    ("Pojištění + ESTA + vízum", 120, ""),
    ("Rezerva / drobné (10%)", 0, "Spočítáno níže"),
]

subtotal = sum(b[1] for b in budget)
reserve = round(subtotal * 0.10)
budget[-1] = ("Rezerva / drobné (10%)", reserve, "Spočítáno automaticky")

grand_total = subtotal + reserve

for cat, price, note in budget:
    pct = round((price / grand_total) * 100, 1)
    ws7.append([cat, price, f"{pct} %", note])

ws7.append(["CELKEM (1 osoba)", grand_total, "100 %", ""])
last = ws7.max_row
for c in range(1, 5):
    ws7.cell(row=last, column=c).fill = TOTAL_FILL
    ws7.cell(row=last, column=c).font = Font(bold=True, size=12, color="C00000")

ws7.append([])
ws7.append(["Pro 2 osoby (společný pokoj, sdílené auto):",
            grand_total * 2 - total_hotels - total_car // 2, "—", "Hotely a auto se nedublují"])
ws7.cell(row=ws7.max_row, column=1).font = Font(bold=True)
ws7.cell(row=ws7.max_row, column=2).font = Font(bold=True, color="C00000")

for r in range(2, ws7.max_row + 1):
    for c in range(1, 5):
        ws7.cell(row=r, column=c).alignment = WRAP
        ws7.cell(row=r, column=c).border = BORDER

auto_width(ws7, [42, 22, 14, 40])

# ============================================================
# Sheet 8 — Checklist
# ============================================================
ws8 = wb.create_sheet("Checklist")
ws8.append(["Kategorie", "Položka", "Termín", "Hotovo?"])
style_header(ws8, 1, 4)

checklist = [
    ("Dokumenty", "Platnost pasu (min. 6 měsíců po návratu)", "ASAP", ""),
    ("Dokumenty", "ESTA žádost (21 USD, 2 roky platnost)", "Min. 72h před odletem", ""),
    ("Dokumenty", "Cestovní pojištění (limit min. 10 mil. Kč pro USA)", "Před odletem", ""),
    ("Dokumenty", "Mezinárodní řidičák", "Min. 2 týdny předem", ""),
    ("Rezervace", "Letenky (multi-city, 7 segmentů)", "3–6 měsíců předem", ""),
    ("Rezervace", "Pearl Harbor tickets (recreation.gov)", "60 dní předem", ""),
    ("Rezervace", "Haleakala sunrise permit", "60 dní předem", ""),
    ("Rezervace", "Hanauma Bay (2 dny předem online)", "2 dny předem", ""),
    ("Rezervace", "Diamond Head reservation", "14 dní předem", ""),
    ("Rezervace", "Old Lahaina Lu'au", "30+ dní předem", ""),
    ("Rezervace", "Molokini Crater snorkel cruise", "30+ dní předem", ""),
    ("Rezervace", "Mauna Kea Summit tour", "14 dní předem", ""),
    ("Rezervace", "Na Pali Coast boat tour", "30+ dní předem", ""),
    ("Rezervace", "Kauai helikoptéra (Blue Hawaiian / Jack Harter)", "14 dní předem", ""),
    ("Rezervace", "Hotely (8 destinací)", "3+ měsíce předem", ""),
    ("Rezervace", "Půjčovny aut (Maui, Big Island, Kauai)", "1+ měsíc předem", ""),
    ("Finance", "Notifikace banky o cestě do USA", "Týden předem", ""),
    ("Finance", "Revolut / Wise pro USD platby (bez fee)", "Týden předem", ""),
    ("Finance", "Trochu USD cash (200 USD na tipy)", "Týden předem", ""),
    ("Telefon", "eSIM USA (Airalo, Holafly) — 4G data", "Den před odletem", ""),
    ("Balení", "Plavky, šnorchl (pro Hanauma Bay vlastní)", "—", ""),
    ("Balení", "Pohorky (Diamond Head, Kilauea Iki)", "—", ""),
    ("Balení", "Teplá bunda (Haleakala -5°C, Mauna Kea -10°C!)", "—", ""),
    ("Balení", "Reef-safe sunscreen (na Havaji povinné)", "—", ""),
    ("Balení", "Univerzální adaptér (US plug A/B)", "—", ""),
    ("Balení", "Power banka pro celodenní výlety", "—", ""),
]
for c in checklist:
    ws8.append(c)

for r in range(2, ws8.max_row + 1):
    for col in range(1, 5):
        ws8.cell(row=r, column=col).alignment = WRAP
        ws8.cell(row=r, column=col).border = BORDER

auto_width(ws8, [16, 50, 22, 12])

# ============================================================
# Sheet 9 — Time zones
# ============================================================
ws9 = wb.create_sheet("Časové pásma")
ws9.append(["Místo", "UTC offset", "Rozdíl vs. Praha", "Pozn."])
style_header(ws9, 1, 4)
tz = [
    ("Praha (CET/CEST)", "+1 / +2", "0 h", "Výchozí"),
    ("Los Angeles (PST/PDT)", "-8 / -7", "-9 h", "Jet lag silný, ale směr OK"),
    ("Honolulu / Maui / Big Island / Kauai (HST)", "-10", "-11 h / -12 h", "Bez letního času! V létě -12 h"),
    ("Miami (EST/EDT)", "-5 / -4", "-6 h", "Po Havaji +5/6 h vpřed → nejtěžší jet lag"),
    ("Návrat Praha", "+1 / +2", "0", "Po MIA jednou nocí v letadle = OK"),
]
for t in tz:
    ws9.append(t)
for r in range(2, ws9.max_row + 1):
    for c in range(1, 5):
        ws9.cell(row=r, column=c).alignment = WRAP
        ws9.cell(row=r, column=c).border = BORDER
auto_width(ws9, [42, 14, 22, 42])

# Save
out = "/Users/jiri.charousek/Projects/havaj-trip-2027/itinerar_havaj_2027.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"  Letenky:    {total_flights:>5} EUR")
print(f"  Hotely:     {total_hotels:>5} EUR")
print(f"  Auta:       {total_car:>5} EUR")
print(f"  Aktivity:   {total_acts:>5} EUR")
print(f"  Jídlo:      {food_total:>5} EUR")
print(f"  CELKEM 1os: {grand_total:>5} EUR")
